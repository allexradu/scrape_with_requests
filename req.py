import requests
from datetime import datetime
import json
import os
import openpyxl as xl
from get_product_siemens import get_data

json_file = os.path.join(os.getcwd(), 'json', 'table.json')
excel_file_path = os.path.join(os.getcwd(), 'excel', 'a.xlsx')

wb = xl.load_workbook(excel_file_path)
sh = wb[wb.sheetnames[0]]

current_col = 2
time = datetime.now()
total_time = datetime.now()

for row in range(1, sh.max_row + 1):
    code = sh.cell(row, 1).value
    data = get_data(code)

    current_time = datetime.now() - time
    print(f'row {row}/{sh.max_row} time {current_time}, total time = {datetime.now() - total_time}')
    time = datetime.now()
    if data is not None:
        if data != {}:
            for key in data.keys():
                sh.cell(row, current_col).value = key
                sh.cell(row, current_col + 1).value = data[key]
                current_col += 2
        else:
            sh.cell(row, current_col).value = 'n/a1'
            sh.cell(row, current_col + 1).value = 'n/a1'
            current_col += 2
    else:
        sh.cell(row, current_col).value = 'n/a2'
        sh.cell(row, current_col + 1).value = 'n/a2'
        current_col += 2
    current_col = 2

wb.save(excel_file_path)
